from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date
from app import db
from app.models.student import Student
from app.models.attendance import Attendance
from app.models.attendance_log import AttendanceLog
from app.models.class_room import ClassRoom
from app.utils.constants import (
    EXCEL_HEADER_COLOR, EXCEL_STATUS_COLORS, 
    ERROR_MESSAGES, ATTENDANCE_STATUSES
)
from app.utils.helpers import ensure_upload_directories, get_upload_path
import io
import os
import logging

logger = logging.getLogger(__name__)


class ExcelExportService:
    
    HEADER_FILL = PatternFill(start_color=EXCEL_HEADER_COLOR.replace('#', ''), 
                             end_color=EXCEL_HEADER_COLOR.replace('#', ''), 
                             fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    @staticmethod
    def export_students_to_excel(classroom_id=None):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Students"
            
            headers = ['STT', 'Mã HS', 'Họ Tên', 'Giới Tính', 'Ngày Sinh', 
                      'Địa Chỉ', 'SĐT', 'Nhận diện khuôn mặt', 'Ghi Chú']
            
            # Apply headers with styling
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = ExcelExportService.HEADER_FONT
                cell.fill = ExcelExportService.HEADER_FILL
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = ExcelExportService.BORDER
            
            # Get students data
            query = db.session.query(Student).filter_by(is_active=True)
            if classroom_id:
                query = query.filter_by(classroom_id=classroom_id)
            
            students = query.order_by(Student.student_code).all()
            
            # Fill data rows
            for idx, student in enumerate(students, 2):
                ws.cell(row=idx, column=1).value = idx - 1
                ws.cell(row=idx, column=2).value = student.student_code
                ws.cell(row=idx, column=3).value = student.full_name
                ws.cell(row=idx, column=4).value = student.gender
                ws.cell(row=idx, column=5).value = student.date_of_birth
                ws.cell(row=idx, column=6).value = student.address
                ws.cell(row=idx, column=7).value = student.phone
                ws.cell(row=idx, column=8).value = 'Đã sẵn sàng' if student.face_recognition_enabled else 'Chưa sẵn sàng'
                
                # Apply borders to data cells
                for col in range(1, len(headers) + 1):
                    ws.cell(row=idx, column=col).border = ExcelExportService.BORDER
            
            # Auto-fit columns
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 25
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 15
            
            logger.info(f'Students exported to Excel: {len(students)} records')
            return wb
            
        except Exception as e:
            logger.error(f'Error exporting students to Excel: {str(e)}')
            raise
    
    @staticmethod
    def export_attendance_report(attendance_log_id=None, start_date=None, end_date=None, classroom_id=None):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        
        headers = ['STT', 'Mã HS', 'Họ Tên', 'Ngày', 'Buổi', 'Trạng Thái', 'Giờ Điểm Danh', 'Ghi Chú']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = ExcelExportService.HEADER_FONT
            cell.fill = ExcelExportService.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        query = db.session.query(Attendance).join(AttendanceLog)
        
        if attendance_log_id:
            query = query.filter(Attendance.attendance_log_id == attendance_log_id)
        
        if classroom_id:
            query = query.filter(Attendance.classroom_id == classroom_id)
        
        if start_date and end_date:
            query = query.filter(AttendanceLog.session_date.between(start_date, end_date))
        
        records = query.order_by(AttendanceLog.session_date.desc(), 
                                Attendance.student_id).all()
        
        status_colors = {
            'present': 'C6EFCE',
            'absent': 'FFC7CE',
            'late': 'FFEB9C',
            'excused': 'B4C7E7'
        }
        
        for idx, record in enumerate(records, 2):
            ws.cell(row=idx, column=1).value = idx - 1
            ws.cell(row=idx, column=2).value = record.student.student_code
            ws.cell(row=idx, column=3).value = record.student.full_name
            ws.cell(row=idx, column=4).value = record.attendance_log.session_date
            ws.cell(row=idx, column=5).value = record.attendance_log.session_type
            ws.cell(row=idx, column=6).value = record.status.capitalize()
            ws.cell(row=idx, column=7).value = record.check_in_time.strftime('%H:%M:%S') if record.check_in_time else '-'
            ws.cell(row=idx, column=8).value = record.notes or ''
            
            fill_color = status_colors.get(record.status, "FFFFFF")
            for col in range(1, 9):
                ws.cell(row=idx, column=col).fill = PatternFill(start_color=fill_color, 
                                                                 end_color=fill_color, 
                                                                 fill_type="solid")
        
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 20
        
        return wb
    
    @staticmethod
    def export_classroom_summary(classroom_id, start_date=None, end_date=None):
        classroom = db.session.query(ClassRoom).get(classroom_id)
        if not classroom:
            raise ValueError("Classroom not found")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        
        ws['A1'] = f"Báo Cáo Điểm Danh - {classroom.class_name}"
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A2'] = f"Thời gian: {start_date or 'All'} đến {end_date or 'All'}"
        ws['A2'].font = Font(size=11)
        
        headers = ['STT', 'Mã HS', 'Họ Tên', 'Tổng Buổi', 'Có Mặt', 'Vắng', 'Muộn', 'Có Phép', 'Tỷ Lệ (%)']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = ExcelExportService.HEADER_FONT
            cell.fill = ExcelExportService.HEADER_FILL
        
        students = db.session.query(Student).filter_by(
            classroom_id=classroom_id,
            is_active=True
        ).all()
        
        for idx, student in enumerate(students, 5):
            query = db.session.query(Attendance).filter_by(student_id=student.id)
            
            if start_date and end_date:
                query = query.join(AttendanceLog).filter(
                    AttendanceLog.session_date.between(start_date, end_date)
                )
            
            records = query.all()
            
            total = len(records)
            present = sum(1 for r in records if r.status == 'present')
            absent = sum(1 for r in records if r.status == 'absent')
            late = sum(1 for r in records if r.status == 'late')
            excused = sum(1 for r in records if r.status == 'excused')
            
            rate = (present / total * 100) if total > 0 else 0
            
            ws.cell(row=idx, column=1).value = idx - 4
            ws.cell(row=idx, column=2).value = student.student_code
            ws.cell(row=idx, column=3).value = student.full_name
            ws.cell(row=idx, column=4).value = total
            ws.cell(row=idx, column=5).value = present
            ws.cell(row=idx, column=6).value = absent
            ws.cell(row=idx, column=7).value = late
            ws.cell(row=idx, column=8).value = excused
            ws.cell(row=idx, column=9).value = round(rate, 2)
            ws.cell(row=idx, column=9).number_format = '0.00'
        
        for col in range(1, 10):
            ws.column_dimensions[chr(64 + col)].width = 12
        
        return wb
    
    @staticmethod
    def export_to_bytes(workbook):
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
